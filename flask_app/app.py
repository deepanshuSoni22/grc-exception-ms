import sys
import os

# Ensure local directories are in the python path for imports
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
sys.path.append(os.path.join(os.path.dirname(__file__), '..', 'engine'))

from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, session
from database import get_db, init_db
from datetime import date, datetime, timedelta
from scheduler import start_scheduler
from functools import wraps
import hashlib

from risk_engine import calculate_risk_score, detect_anomaly
from email_notifications import send_audit_report_email

app = Flask(__name__)
app.secret_key = "grc_secret_2026"

# Initialize database
init_db()

# Start scheduler only in the main worker process
if not app.debug or os.environ.get("WERKZEUG_RUN_MAIN") == "true":
    start_scheduler()

TODAY = date(2026, 4, 15)

EXCEPTION_TYPES = [
    "Admin Access", "Root Access", "Firewall Exception",
    "Encryption Waiver", "VPN Bypass", "Data Retention Waiver",
    "Password Policy Exception", "MFA Bypass", "Network Access"
]

BUSINESS_UNITS = [
    "Finance", "HR", "Engineering", "Sales",
    "Legal", "Operations", "IT", "Marketing"
]

RISK_LEVELS = ["LOW", "MEDIUM", "HIGH", "CRITICAL"]


# ── Auth decorator ────────────────────────────────────────────────
def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if "user" not in session:
            flash("Please log in to access this page.", "error")
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated


# ── Helper ────────────────────────────────────────────────────────
def row_to_dict(row):
    return dict(row)


def enrich(exc):
    end = datetime.strptime(exc["end_date"], "%Y-%m-%d").date()
    days_to_expiry = (end - TODAY).days

    if exc["status"] == "REVOKED":
        exc["display_status"] = "revoked"
    elif end < TODAY and exc["status"] == "ACTIVE":
        exc["display_status"] = "expired"
    elif days_to_expiry <= 30 and exc["status"] == "ACTIVE":
        exc["display_status"] = "expiring"
    elif exc["status"] == "PENDING":
        exc["display_status"] = "pending"
    else:
        exc["display_status"] = "active"

    exc["days_to_expiry"] = days_to_expiry
    return exc


def gather_report_stats():
    db = get_db()
    all_exc = [row_to_dict(r) for r in db.execute("SELECT * FROM exceptions").fetchall()]
    db.close()

    active_exc = [e for e in all_exc if e["status"] == "ACTIVE"]

    expired_not_revoked = [
        e for e in active_exc
        if datetime.strptime(e["end_date"], "%Y-%m-%d").date() < TODAY
    ]
    expiring_this_month = [
        e for e in active_exc
        if 0 <= (datetime.strptime(e["end_date"], "%Y-%m-%d").date() - TODAY).days <= 30
    ]
    overdue_review = [
        e for e in all_exc
        if e["status"] == "PENDING"
        and (TODAY - datetime.strptime(e["start_date"], "%Y-%m-%d").date()).days > 30
    ]

    return {
        "report_date": TODAY.strftime("%Y-%m-%d"),
        "total_active": len(active_exc),
        "expiring_this_month": expiring_this_month,
        "expired_not_revoked": expired_not_revoked,
        "overdue_review": overdue_review,
    }


# ── LOGIN ─────────────────────────────────────────────────────────
@app.route("/login", methods=["GET", "POST"])
def login():
    if "user" in session:
        return redirect(url_for("dashboard"))

    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")
        pw_hash  = hashlib.sha256(password.encode()).hexdigest()

        db = get_db()
        user = db.execute(
            "SELECT * FROM users WHERE username=? AND password_hash=?",
            (username, pw_hash)
        ).fetchone()
        db.close()

        if user:
            session["user"]     = username
            session["role"]     = user["role"]
            flash(f"Welcome back, {username}!", "success")
            return redirect(url_for("dashboard"))
        else:
            flash("Invalid username or password.", "error")

    return render_template("login.html")


# ── LOGOUT ────────────────────────────────────────────────────────
@app.route("/logout")
def logout():
    session.clear()
    flash("You have been logged out.", "success")
    return redirect(url_for("login"))


# ── HOME ──────────────────────────────────────────────────────────
@app.route("/")
@login_required
def index():
    return redirect(url_for("dashboard"))


# ── DASHBOARD ─────────────────────────────────────────────────────
@app.route("/dashboard")
@login_required
def dashboard():
    db = get_db()

    total   = db.execute("SELECT COUNT(*) FROM exceptions").fetchone()[0]
    active  = db.execute("SELECT COUNT(*) FROM exceptions WHERE status='ACTIVE'").fetchone()[0]
    pending = db.execute("SELECT COUNT(*) FROM exceptions WHERE status='PENDING'").fetchone()[0]
    revoked = db.execute("SELECT COUNT(*) FROM exceptions WHERE status='REVOKED'").fetchone()[0]

    all_active = db.execute("SELECT * FROM exceptions WHERE status='ACTIVE'").fetchall()
    expired_not_revoked = sum(
        1 for r in all_active
        if datetime.strptime(r["end_date"], "%Y-%m-%d").date() < TODAY
    )
    expiring_soon = sum(
        1 for r in all_active
        if 0 <= (datetime.strptime(r["end_date"], "%Y-%m-%d").date() - TODAY).days <= 30
    )

    critical = db.execute(
        "SELECT COUNT(*) FROM exceptions WHERE risk_level='CRITICAL' AND status='ACTIVE'"
    ).fetchone()[0]
    high = db.execute(
        "SELECT COUNT(*) FROM exceptions WHERE risk_level='HIGH' AND status='ACTIVE'"
    ).fetchone()[0]

    risk_dist = db.execute(
        "SELECT risk_level, COUNT(*) as cnt FROM exceptions WHERE status='ACTIVE' GROUP BY risk_level"
    ).fetchall()
    risk_chart = {r["risk_level"]: r["cnt"] for r in risk_dist}

    type_dist = db.execute(
        "SELECT type, COUNT(*) as cnt FROM exceptions WHERE status='ACTIVE' GROUP BY type ORDER BY cnt DESC LIMIT 6"
    ).fetchall()
    type_chart = {"labels": [r["type"] for r in type_dist],
                  "data":   [r["cnt"]  for r in type_dist]}

    top_risk = db.execute(
        "SELECT * FROM exceptions WHERE status='ACTIVE' ORDER BY risk_score DESC LIMIT 5"
    ).fetchall()
    top_risk = [enrich(row_to_dict(r)) for r in top_risk]

    anomalies = db.execute(
        "SELECT * FROM exceptions WHERE predicted_anomaly=1 ORDER BY risk_score DESC LIMIT 5"
    ).fetchall()
    anomalies = [enrich(row_to_dict(r)) for r in anomalies]

    top_requesters = db.execute("""
        SELECT requester,
               COUNT(*) as total_exceptions,
               SUM(risk_score) as total_risk,
               MAX(risk_score) as max_risk,
               SUM(CASE WHEN risk_level IN ('HIGH','CRITICAL') THEN 1 ELSE 0 END) as high_count
        FROM exceptions
        WHERE status='ACTIVE'
        GROUP BY requester
        ORDER BY total_risk DESC
        LIMIT 8
    """).fetchall()
    top_requesters = [dict(r) for r in top_requesters]

    db.close()

    return render_template("dashboard.html",
        total=total, active=active, pending=pending,
        revoked=revoked, expired_not_revoked=expired_not_revoked,
        expiring_soon=expiring_soon, critical=critical, high=high,
        risk_chart=risk_chart, type_chart=type_chart,
        top_risk=top_risk, anomalies=anomalies,
        top_requesters=top_requesters, today=TODAY
    )


# ── EXCEPTIONS TABLE ──────────────────────────────────────────────
@app.route("/exceptions")
@login_required
def exceptions():
    db = get_db()

    status_filter = request.args.get("status", "")
    risk_filter   = request.args.get("risk_level", "")
    type_filter   = request.args.get("type", "")
    search        = request.args.get("search", "")
    sort_by       = request.args.get("sort", "risk_score")
    sort_dir      = request.args.get("dir", "desc")

    query  = "SELECT * FROM exceptions WHERE 1=1"
    params = []

    if status_filter:
        query += " AND status=?"; params.append(status_filter)
    if risk_filter:
        query += " AND risk_level=?"; params.append(risk_filter)
    if type_filter:
        query += " AND type=?"; params.append(type_filter)
    if search:
        query += " AND (exception_id LIKE ? OR requester LIKE ? OR justification LIKE ?)"
        params += [f"%{search}%", f"%{search}%", f"%{search}%"]

    allowed_sorts = ["exception_id","type","requester","risk_score","risk_level",
                     "status","start_date","end_date","business_unit"]
    if sort_by not in allowed_sorts:
        sort_by = "risk_score"
    direction = "DESC" if sort_dir == "desc" else "ASC"
    query += f" ORDER BY {sort_by} {direction}"

    rows = db.execute(query, params).fetchall()
    rows = [enrich(row_to_dict(r)) for r in rows]

    exc_types = db.execute("SELECT DISTINCT type FROM exceptions ORDER BY type").fetchall()
    db.close()

    return render_template("exceptions.html",
        exceptions=rows, exc_types=[r["type"] for r in exc_types],
        status_filter=status_filter, risk_filter=risk_filter,
        type_filter=type_filter, search=search,
        sort_by=sort_by, sort_dir=sort_dir,
        risk_levels=RISK_LEVELS
    )


# ── ADD EXCEPTION ─────────────────────────────────────────────────
@app.route("/add", methods=["GET", "POST"])
@login_required
def add_exception():
    if request.method == "POST":
        db = get_db()

        last = db.execute(
            "SELECT exception_id FROM exceptions WHERE exception_id LIKE 'EXC-%' ORDER BY id DESC LIMIT 1"
        ).fetchone()
        if last:
            try:
                last_num = int(last["exception_id"].split("-")[1])
            except Exception:
                last_num = 600
            new_id = f"EXC-{last_num+1:04d}"
            while db.execute("SELECT 1 FROM exceptions WHERE exception_id=?", (new_id,)).fetchone():
                last_num += 1
                new_id = f"EXC-{last_num+1:04d}"
        else:
            new_id = "EXC-0601"

        data = {
            "exception_id":  new_id,
            "type":          request.form["type"],
            "requester":     request.form["requester"],
            "approver":      request.form["approver"],
            "business_unit": request.form["business_unit"],
            "justification": request.form["justification"],
            "start_date":    request.form["start_date"],
            "end_date":      request.form["end_date"],
            "status":        request.form["status"],
            "risk_level":    request.form["risk_level"],
        }

        score = calculate_risk_score(data)
        is_anomaly, pred_type, pred_severity = detect_anomaly(data)

        db.execute("""
            INSERT INTO exceptions
            (exception_id, type, requester, approver, business_unit,
             justification, start_date, end_date, status, risk_level,
             risk_score, predicted_anomaly, predicted_type, predicted_severity,
             renewal_count)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, (
            data["exception_id"], data["type"], data["requester"],
            data["approver"], data["business_unit"], data["justification"],
            data["start_date"], data["end_date"], data["status"], data["risk_level"],
            score, int(is_anomaly), pred_type, pred_severity, 0
        ))
        db.commit()
        db.close()

        flash(f"Exception {new_id} logged successfully! Risk Score: {score}/100", "success")
        return redirect(url_for("exceptions"))

    return render_template("add_exception.html",
        exception_types=EXCEPTION_TYPES,
        business_units=BUSINESS_UNITS,
        risk_levels=RISK_LEVELS,
        today=TODAY.isoformat()
    )


# ── ALERTS ────────────────────────────────────────────────────────
@app.route("/alerts")
@login_required
def alerts():
    db = get_db()
    all_exc = [row_to_dict(r) for r in db.execute("SELECT * FROM exceptions").fetchall()]
    db.close()

    expired_active, expiring_soon, long_running, stalled, critical_list = [], [], [], [], []

    for exc in all_exc:
        end   = datetime.strptime(exc["end_date"],   "%Y-%m-%d").date()
        start = datetime.strptime(exc["start_date"], "%Y-%m-%d").date()
        days_running      = (TODAY - start).days
        days_to_expiry    = (end - TODAY).days
        days_since_expiry = (TODAY - end).days

        exc = enrich(exc)

        if exc["status"] == "ACTIVE" and end < TODAY:
            exc["alert_msg"] = f"Expired {days_since_expiry} days ago — NOT revoked!"
            expired_active.append(exc)
        elif exc["status"] == "ACTIVE" and 0 <= days_to_expiry <= 30:
            exc["alert_msg"] = f"Expires in {days_to_expiry} days"
            expiring_soon.append(exc)
        elif days_running > 180 and exc["status"] == "ACTIVE":
            exc["alert_msg"] = f"Running for {days_running} days without renewal"
            long_running.append(exc)
        elif exc["status"] == "PENDING" and days_running > 30:
            exc["alert_msg"] = f"Pending review for {days_running} days"
            stalled.append(exc)

        if exc["risk_level"] == "CRITICAL" and exc["status"] == "ACTIVE":
            exc["alert_msg"] = exc.get("alert_msg", "Critical risk — needs immediate review")
            critical_list.append(exc)

    return render_template("alerts.html",
        expired_active=expired_active,
        expiring_soon=expiring_soon,
        long_running=long_running,
        stalled=stalled,
        critical_list=critical_list
    )


# ── AUDIT REPORT ──────────────────────────────────────────────────
@app.route("/report")
@login_required
def report():
    db = get_db()
    all_exc = [row_to_dict(r) for r in db.execute("SELECT * FROM exceptions").fetchall()]
    db.close()

    active_exc = [e for e in all_exc if e["status"] == "ACTIVE"]

    expired_not_revoked = [
        e for e in active_exc
        if datetime.strptime(e["end_date"], "%Y-%m-%d").date() < TODAY
    ]
    expiring_this_month = [
        e for e in active_exc
        if 0 <= (datetime.strptime(e["end_date"], "%Y-%m-%d").date() - TODAY).days <= 30
    ]
    overdue_review = [
        e for e in all_exc
        if e["status"] == "PENDING"
        and (TODAY - datetime.strptime(e["start_date"], "%Y-%m-%d").date()).days > 30
    ]

    high_risk   = [e for e in active_exc if e["risk_level"] in ("HIGH", "CRITICAL")]
    medium_risk = [e for e in active_exc if e["risk_level"] == "MEDIUM"]
    low_risk    = [e for e in active_exc if e["risk_level"] == "LOW"]

    type_counts = {}
    for e in active_exc:
        type_counts[e["type"]] = type_counts.get(e["type"], 0) + 1

    top_high = sorted(
        [e for e in active_exc if e["risk_level"] in ("HIGH", "CRITICAL")],
        key=lambda x: x["risk_score"], reverse=True
    )[:10]
    top_high = [enrich(e) for e in top_high]

    with_approver = sum(1 for e in all_exc if e.get("approver"))
    approval_pct  = round(with_approver / len(all_exc) * 100) if all_exc else 0

    report_date = TODAY.strftime("%Y-%m-%d")
    time_range  = f"{(TODAY - timedelta(days=90)).strftime('%Y-%m-%d')} to {report_date}"

    return render_template("report.html",
        report_date=report_date,
        time_range=time_range,
        total_active=len(active_exc),
        high_risk_count=len(high_risk),
        medium_risk_count=len(medium_risk),
        low_risk_count=len(low_risk),
        expiring_this_month=expiring_this_month,
        expired_not_revoked=expired_not_revoked,
        overdue_review=overdue_review,
        type_counts=type_counts,
        top_high=top_high,
        approval_pct=approval_pct,
        total_exceptions=len(all_exc)
    )


# ── SEND REPORT VIA EMAIL ─────────────────────────────────────────
@app.route("/send-report", methods=["POST"])
@login_required
def send_report():
    stats = gather_report_stats()
    success, message = send_audit_report_email(stats)
    flash(message, "success" if success else "error")
    return redirect(url_for("report"))


# ── EXCEL EXPORT ──────────────────────────────────────────────────
@app.route("/export/excel")
@login_required
def export_excel():
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    from flask import send_file
    import io

    db   = get_db()
    rows = [row_to_dict(r) for r in db.execute("SELECT * FROM exceptions ORDER BY risk_score DESC").fetchall()]
    db.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Exception Registry"

    headers = ["Exception ID","Type","Requester","Approver","Business Unit",
               "Justification","Start Date","End Date","Status",
               "Risk Level","Risk Score","Anomaly","Anomaly Type","Renewal Count"]

    header_fill = PatternFill("solid", fgColor="1a1a2e")
    header_font = Font(color="FFFFFF", bold=True)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    fills = {
        "ACTIVE":  PatternFill("solid", fgColor="d4edda"),
        "PENDING": PatternFill("solid", fgColor="fff3cd"),
        "REVOKED": PatternFill("solid", fgColor="f8f9fa"),
        "EXPIRED": PatternFill("solid", fgColor="f8d7da"),
    }

    for row_num, exc in enumerate(rows, 2):
        end = datetime.strptime(exc["end_date"], "%Y-%m-%d").date()
        display  = "EXPIRED" if (end < TODAY and exc["status"] == "ACTIVE") else exc["status"]
        row_fill = fills.get(display, fills["ACTIVE"])

        values = [
            exc["exception_id"], exc["type"], exc["requester"], exc["approver"],
            exc["business_unit"], exc["justification"], exc["start_date"],
            exc["end_date"], exc["status"], exc["risk_level"],
            exc["risk_score"],
            "YES" if exc["predicted_anomaly"] else "NO",
            exc["predicted_type"],
            exc.get("renewal_count", 0)
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col, value=val)
            cell.fill = row_fill

    for col in ws.columns:
        max_len = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"grc_exceptions_{TODAY}.xlsx"
    )


# ── EXCEPTION DETAIL VIEW ─────────────────────────────────────────
@app.route("/exception/<exception_id>")
@login_required
def exception_detail(exception_id):
    db = get_db()
    exc = db.execute(
        "SELECT * FROM exceptions WHERE exception_id=?", (exception_id,)
    ).fetchone()
    if not exc:
        db.close()
        return "Exception not found", 404

    exc = enrich(row_to_dict(exc))

    start = datetime.strptime(exc["start_date"], "%Y-%m-%d").date()
    end   = datetime.strptime(exc["end_date"],   "%Y-%m-%d").date()
    days_running      = (TODAY - start).days
    days_since_expiry = (TODAY - end).days
    days_to_expiry    = (end - TODAY).days

    alerts = []
    recommendation = ""

    if end < TODAY and exc["status"] == "ACTIVE":
        alerts.append(f"EXPIRED_NOT_REVOKED: End date passed {days_since_expiry} days ago; still marked ACTIVE")
        recommendation = f"REVOKE IMMEDIATELY — exception expired {days_since_expiry} days ago and was never closed"

    if exc["risk_level"] == "CRITICAL" and exc["status"] == "ACTIVE":
        alerts.append("CRITICAL_RISK: Marked CRITICAL risk level — requires re-review by CISO")
        if not recommendation:
            recommendation = "ESCALATE TO CISO — critical risk exception requires immediate sign-off"

    if exc["risk_level"] == "HIGH" and days_running > 90 and exc["status"] == "ACTIVE":
        alerts.append(f"HIGH_RISK_LONG_RUNNING: HIGH risk exception active for {days_running} days without renewal review")
        if not recommendation:
            recommendation = "SCHEDULE REVIEW — high risk exception overdue for renewal assessment"

    if days_running > 180 and exc["status"] == "ACTIVE":
        alerts.append(f"LONG_RUNNING: Exception has been active for {days_running} days (threshold: 180 days)")
        if not recommendation:
            recommendation = "RENEWAL REQUIRED — exception exceeds 180-day threshold, must be formally renewed"

    if exc["status"] == "PENDING" and days_running > 30:
        alerts.append(f"STALLED_REVIEW: Pending approval for {days_running} days (limit: 30 days)")
        if not recommendation:
            recommendation = "ESCALATE TO APPROVER — approval has been pending beyond acceptable SLA"

    if exc["type"] in ("Admin Access", "Root Access"):
        alerts.append("ELEVATED_PRIVILEGE: Admin/Root access should be strictly temporary and time-limited")

    if 0 <= days_to_expiry <= 30 and exc["status"] == "ACTIVE":
        alerts.append(f"EXPIRING_SOON: Exception expires in {days_to_expiry} days — renew or revoke")
        if not recommendation:
            recommendation = f"ACTION NEEDED — renew or revoke before {exc['end_date']}"

    if not alerts:
        alerts.append("No active alerts — exception is within policy parameters")
        recommendation = "No immediate action required — continue standard monitoring"

    label = db.execute(
        "SELECT * FROM exception_labels WHERE exception_id=?", (exception_id,)
    ).fetchone()
    label = dict(label) if label else None

    db.close()

    return render_template("exception_detail.html",
        exc=exc, alerts=alerts, recommendation=recommendation,
        days_running=days_running, days_to_expiry=days_to_expiry, label=label
    )


# ── REVOKE EXCEPTION ──────────────────────────────────────────────
@app.route("/revoke/<exception_id>", methods=["POST"])
@login_required
def revoke_exception(exception_id):
    db = get_db()
    db.execute("UPDATE exceptions SET status='REVOKED' WHERE exception_id=?", (exception_id,))
    db.commit()
    db.close()
    flash(f"Exception {exception_id} has been revoked.", "success")
    return redirect(url_for("exceptions"))


# ── RENEW EXCEPTION ───────────────────────────────────────────────
@app.route("/renew/<exception_id>", methods=["POST"])
@login_required
def renew_exception(exception_id):
    db = get_db()
    exc = row_to_dict(db.execute(
        "SELECT * FROM exceptions WHERE exception_id=?", (exception_id,)
    ).fetchone())

    new_end = (TODAY + timedelta(days=90)).isoformat()
    exc["end_date"] = new_end
    new_score = calculate_risk_score(exc)
    is_anomaly, pred_type, pred_severity = detect_anomaly(exc)

    db.execute("""
        UPDATE exceptions
        SET end_date=?, status='ACTIVE', risk_score=?,
            predicted_anomaly=?, predicted_type=?, predicted_severity=?,
            renewal_count = COALESCE(renewal_count, 0) + 1
        WHERE exception_id=?
    """, (new_end, new_score, int(is_anomaly), pred_type, pred_severity, exception_id))
    db.commit()
    db.close()
    flash(f"Exception {exception_id} renewed until {new_end}. New risk score: {new_score}/100", "success")
    return redirect(url_for("exceptions"))


# ── NOTIFICATION LOG ──────────────────────────────────────────────
@app.route("/notification-log")
@login_required
def notification_log():
    log_path = os.path.join(os.path.dirname(__file__), 'notification_log.txt')
    entries = []
    if os.path.exists(log_path):
        with open(log_path, 'r') as f:
            for line in f.readlines():
                line = line.strip()
                if line:
                    entries.append(line)
    entries.reverse()
    return render_template("notification_log.html", entries=entries)


# ── EVALUATION PAGE ───────────────────────────────────────────────
@app.route("/evaluation")
@login_required
def evaluation():
    from sklearn.metrics import classification_report, confusion_matrix

    db = get_db()
    rows = db.execute("""
        SELECT e.predicted_anomaly, e.predicted_type,
               l.is_anomaly, l.anomaly_type
        FROM exceptions e
        JOIN exception_labels l ON e.exception_id = l.exception_id
    """).fetchall()
    db.close()

    if not rows:
        return render_template("evaluation.html", error="No label data found.")

    y_true = [bool(r["is_anomaly"])        for r in rows]
    y_pred = [bool(r["predicted_anomaly"]) for r in rows]

    report = classification_report(y_true, y_pred,
        target_names=["Compliant", "At-Risk Exception"], output_dict=True)

    cm = confusion_matrix(y_true, y_pred).tolist()

    true_types = [r["anomaly_type"]   for r in rows]
    pred_types = [r["predicted_type"] for r in rows]
    type_report = classification_report(true_types, pred_types,
        zero_division=0, output_dict=True)

    critical_total  = sum(1 for r in rows if r["anomaly_type"] in ("EXPIRED_ACTIVE_EXCEPTION","CRITICAL_RISK_EXCEPTION"))
    critical_caught = sum(1 for r in rows if r["anomaly_type"] in ("EXPIRED_ACTIVE_EXCEPTION","CRITICAL_RISK_EXCEPTION") and r["predicted_anomaly"])
    critical_rate   = round(critical_caught / critical_total * 100, 1) if critical_total else 0

    return render_template("evaluation.html",
        report=report, cm=cm,
        type_report=type_report,
        critical_rate=critical_rate,
        critical_caught=critical_caught,
        critical_total=critical_total,
        accuracy=round(report["accuracy"] * 100, 1)
    )


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=int(os.environ.get("PORT", 5000)), debug=True)